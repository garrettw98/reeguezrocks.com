from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2B5797')
alt_fill = PatternFill('solid', fgColor='F2F2F2')
input_font = Font(color='0000FF')
currency_format = '"$"#,##0'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: Funding Tiers (Scaled to 1,500)
ws1 = wb.active
ws1.title = "Funding Tiers"

ws1['A1'] = 'REEGUEZ ROCKS 2026 - FUNDING TIERS (Scaled to 1,500 Attendees)'
ws1['A1'].font = Font(bold=True, size=14)
ws1.merge_cells('A1:F1')

# Phase 1 Header
ws1['A3'] = 'PHASE 1: FOUNDATION (150-200 Attendees)'
ws1['A3'].font = Font(bold=True, size=11, color='2B5797')
ws1.merge_cells('A3:F3')

headers = ['Tier', 'Name', 'Goal', 'Cumulative', 'Category', 'What It Unlocks']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

tiers_phase1 = [
    (1, 'Make It Happen', 3000, 'Infrastructure', 'Insurance, fuel, U-Hauls, contingency'),
    (2, 'First Headliner', 5000, 'Artists', 'Book a name, announce on flyer'),
    (3, 'Basic Production', 3900, 'Production', 'Subphonic ($3,000) + Keith dome ($900)'),
]

row = 5
cumulative = 0
for tier, name, goal, category, desc in tiers_phase1:
    cumulative += goal
    ws1.cell(row=row, column=1, value=tier).border = thin_border
    ws1.cell(row=row, column=2, value=name).border = thin_border
    c = ws1.cell(row=row, column=3, value=goal)
    c.number_format = currency_format
    c.border = thin_border
    c = ws1.cell(row=row, column=4, value=cumulative)
    c.number_format = currency_format
    c.border = thin_border
    ws1.cell(row=row, column=5, value=category).border = thin_border
    ws1.cell(row=row, column=6, value=desc).border = thin_border
    row += 1

# Phase 2 Header
row += 1
ws1.cell(row=row, column=1, value='PHASE 2: GROWTH (300-500 Attendees)').font = Font(bold=True, size=11, color='2B5797')
ws1.merge_cells(f'A{row}:F{row}')
row += 1

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
row += 1

tiers_phase2 = [
    (4, 'Subphonic Upgraded', 3000, 'Production', '+$3k, now $6k total'),
    (5, 'Festival Ambiance', 3000, 'Experience', 'Lighting, projectors, decor'),
    (6, 'Second Headliner', 5000, 'Artists', 'Another name act for diversity'),
    (7, 'Daytime Programming', 2500, 'Experience', 'Yoga, sound baths, workshops'),
]

for tier, name, goal, category, desc in tiers_phase2:
    cumulative += goal
    ws1.cell(row=row, column=1, value=tier).border = thin_border
    ws1.cell(row=row, column=2, value=name).border = thin_border
    c = ws1.cell(row=row, column=3, value=goal)
    c.number_format = currency_format
    c.border = thin_border
    c = ws1.cell(row=row, column=4, value=cumulative)
    c.number_format = currency_format
    c.border = thin_border
    ws1.cell(row=row, column=5, value=category).border = thin_border
    ws1.cell(row=row, column=6, value=desc).border = thin_border
    row += 1

# Phase 3 Header
row += 1
ws1.cell(row=row, column=1, value='PHASE 3: EXPANSION (500-800 Attendees)').font = Font(bold=True, size=11, color='2B5797')
ws1.merge_cells(f'A{row}:F{row}')
row += 1

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
row += 1

tiers_phase3 = [
    (8, 'Subphonic Premium', 3000, 'Production', '+$3k, now $9k total'),
    (9, 'Third Headliner + Support', 5000, 'Artists', 'Fill out the lineup'),
    (10, 'Main Stage Visuals', 5000, 'Production', 'LED wall, expanded lighting'),
    (11, 'Art + Infrastructure', 5000, 'Experience', 'Installations, fencing, signage, power'),
]

for tier, name, goal, category, desc in tiers_phase3:
    cumulative += goal
    ws1.cell(row=row, column=1, value=tier).border = thin_border
    ws1.cell(row=row, column=2, value=name).border = thin_border
    c = ws1.cell(row=row, column=3, value=goal)
    c.number_format = currency_format
    c.border = thin_border
    c = ws1.cell(row=row, column=4, value=cumulative)
    c.number_format = currency_format
    c.border = thin_border
    ws1.cell(row=row, column=5, value=category).border = thin_border
    ws1.cell(row=row, column=6, value=desc).border = thin_border
    row += 1

# Phase 4 Header
row += 1
ws1.cell(row=row, column=1, value='PHASE 4: FESTIVAL SCALE (1,000-1,500 Attendees)').font = Font(bold=True, size=11, color='2B5797')
ws1.merge_cells(f'A{row}:F{row}')
row += 1

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
row += 1

tiers_phase4 = [
    (12, 'Subphonic Festival-Grade', 3000, 'Production', '+$3k, now $12k total'),
    (13, 'Major Headliner', 10000, 'Artists', 'Marquee name to anchor lineup'),
    (14, 'Full Support Lineup', 8000, 'Artists', '4-6 additional mid-tier acts'),
    (15, 'Premium Package', 10000, 'Experience', 'Enhanced PA, aftermovie, medical, security'),
]

for tier, name, goal, category, desc in tiers_phase4:
    cumulative += goal
    ws1.cell(row=row, column=1, value=tier).border = thin_border
    ws1.cell(row=row, column=2, value=name).border = thin_border
    c = ws1.cell(row=row, column=3, value=goal)
    c.number_format = currency_format
    c.border = thin_border
    c = ws1.cell(row=row, column=4, value=cumulative)
    c.number_format = currency_format
    c.border = thin_border
    ws1.cell(row=row, column=5, value=category).border = thin_border
    ws1.cell(row=row, column=6, value=desc).border = thin_border
    row += 1

# Total row
row += 1
ws1.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
c = ws1.cell(row=row, column=4, value=cumulative)
c.number_format = currency_format
c.font = Font(bold=True)

# Column widths
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 45

# Sheet 2: Ticket Pricing
ws2 = wb.create_sheet("Ticket Pricing")

ws2['A1'] = 'TICKET PRICING'
ws2['A1'].font = Font(bold=True, size=14)

headers = ['Ticket Type', 'Price', 'To Venue', 'Margin', 'Music Days']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

tickets = [
    ('4-Day Pass', 99, 40, 59, '4 (Thu-Mon)'),
    ('3-Day Pass', 80, 30, 50, '3 (Fri-Mon)'),
    ('2-Day Pass', 60, 20, 40, '2 (Sat-Sun)'),
    ('1-Day Pass', 30, 10, 20, '1'),
]

for row, (ttype, price, venue, margin, days) in enumerate(tickets, 4):
    ws2.cell(row=row, column=1, value=ttype).border = thin_border
    c = ws2.cell(row=row, column=2, value=price)
    c.number_format = currency_format
    c.border = thin_border
    c = ws2.cell(row=row, column=3, value=venue)
    c.number_format = currency_format
    c.border = thin_border
    c = ws2.cell(row=row, column=4, value=margin)
    c.number_format = currency_format
    c.border = thin_border
    ws2.cell(row=row, column=5, value=days).border = thin_border

# Add-ons
ws2['A10'] = 'ADD-ONS (100% Margin)'
ws2['A10'].font = Font(bold=True)

headers = ['Add-On', 'Price', 'Availability']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=11, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

addons = [
    ('Car Camping', 30, 'Unlimited'),
    ('Cabin (9 beds)', 1000, '6 cabins'),
]

for row, (addon, price, avail) in enumerate(addons, 12):
    ws2.cell(row=row, column=1, value=addon).border = thin_border
    c = ws2.cell(row=row, column=2, value=price)
    c.number_format = currency_format
    c.border = thin_border
    ws2.cell(row=row, column=3, value=avail).border = thin_border

# Meal Plans
ws2['A16'] = 'MEAL PLANS'
ws2['A16'].font = Font(bold=True)

headers = ['Package', 'Meals', 'Price', 'Per Meal']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=17, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

meals = [
    ('Full Festival (Thu-Mon)', 8, 100, 12.50),
    ('Weekend Only (Sat-Mon)', 4, 50, 12.50),
    ('Single Day', 2, 30, 15.00),
]

for row, (pkg, num_meals, price, per) in enumerate(meals, 18):
    ws2.cell(row=row, column=1, value=pkg).border = thin_border
    ws2.cell(row=row, column=2, value=num_meals).border = thin_border
    c = ws2.cell(row=row, column=3, value=price)
    c.number_format = currency_format
    c.border = thin_border
    c = ws2.cell(row=row, column=4, value=per)
    c.number_format = '"$"#,##0.00'
    c.border = thin_border

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 15

# Sheet 3: Revenue Calculator
ws3 = wb.create_sheet("Revenue Calculator")

ws3['A1'] = 'REVENUE CALCULATOR'
ws3['A1'].font = Font(bold=True, size=14)

ws3['A3'] = 'Enter ticket sales below (blue cells):'
ws3['A3'].font = Font(italic=True)

headers = ['Ticket Type', 'Quantity', 'Margin', 'Revenue']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

calc_items = [
    ('4-Day Pass', 59),
    ('3-Day Pass', 50),
    ('2-Day Pass', 40),
    ('1-Day Pass', 20),
    ('Car Camping', 30),
    ('Cabins', 1000),
    ('Full Meal Plan', 25),
]

for row, (item, margin) in enumerate(calc_items, 6):
    ws3.cell(row=row, column=1, value=item).border = thin_border
    qty_cell = ws3.cell(row=row, column=2, value=0)
    qty_cell.font = input_font
    qty_cell.border = thin_border
    c = ws3.cell(row=row, column=3, value=margin)
    c.number_format = currency_format
    c.border = thin_border
    rev_cell = ws3.cell(row=row, column=4, value=f'=B{row}*C{row}')
    rev_cell.number_format = currency_format
    rev_cell.border = thin_border

# Total row
total_row = 13
ws3.cell(row=total_row, column=1, value='TOTAL MARGIN').font = Font(bold=True)
ws3.cell(row=total_row, column=4, value='=SUM(D6:D12)').font = Font(bold=True)
ws3.cell(row=total_row, column=4).number_format = currency_format

# Tier reached
ws3['A15'] = 'TIER REACHED:'
ws3['A15'].font = Font(bold=True)
ws3['B15'] = '=IF(D13>=84400,"Tier 15",IF(D13>=76400,"Tier 14",IF(D13>=68400,"Tier 13",IF(D13>=56400,"Tier 12",IF(D13>=46400,"Tier 11",IF(D13>=41400,"Tier 10",IF(D13>=38400,"Tier 9",IF(D13>=30400,"Tier 8",IF(D13>=25400,"Tier 7",IF(D13>=22900,"Tier 6",IF(D13>=19900,"Tier 5",IF(D13>=14900,"Tier 4",IF(D13>=11900,"Tier 3",IF(D13>=8000,"Tier 2",IF(D13>=3000,"Tier 1","Below Tier 1")))))))))))))))'

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 15

# Sheet 4: Scenarios
ws4 = wb.create_sheet("Scenarios")

ws4['A1'] = 'REVENUE SCENARIOS BY ATTENDANCE'
ws4['A1'].font = Font(bold=True, size=14)

headers = ['Attendance', '4-Day', '3-Day', '2-Day', 'Cabins', 'Camping', 'Meals', 'Total Margin', 'Phase']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

scenarios = [
    (150, 100, 30, 20, 4, 40, 30, 13400, 'Phase 1'),
    (300, 200, 60, 40, 6, 80, 60, 26700, 'Phase 2'),
    (500, 350, 100, 50, 6, 150, 100, 40650, 'Phase 3'),
    (800, 550, 150, 100, 6, 250, 160, 63200, 'Phase 3-4'),
    (1000, 700, 200, 100, 6, 350, 200, 80050, 'Phase 4'),
    (1500, 1000, 300, 200, 6, 500, 300, 110500, 'Phase 4+'),
]

for row, (att, d4, d3, d2, cab, camp, meals, margin, phase) in enumerate(scenarios, 4):
    ws4.cell(row=row, column=1, value=att).border = thin_border
    ws4.cell(row=row, column=2, value=d4).border = thin_border
    ws4.cell(row=row, column=3, value=d3).border = thin_border
    ws4.cell(row=row, column=4, value=d2).border = thin_border
    ws4.cell(row=row, column=5, value=cab).border = thin_border
    ws4.cell(row=row, column=6, value=camp).border = thin_border
    ws4.cell(row=row, column=7, value=meals).border = thin_border
    c = ws4.cell(row=row, column=8, value=margin)
    c.number_format = currency_format
    c.border = thin_border
    ws4.cell(row=row, column=9, value=phase).border = thin_border

ws4['A12'] = 'Formula: (4d×$59) + (3d×$50) + (2d×$40) + (cabins×$1000) + (camp×$30) + (meals×$25)'
ws4['A12'].font = Font(italic=True, size=10)

for col in range(1, 10):
    ws4.column_dimensions[get_column_letter(col)].width = 12

# Sheet 5: Partner Tracker
ws5 = wb.create_sheet("Partners")

ws5['A1'] = 'STAGE PARTNER TRACKER'
ws5['A1'].font = Font(bold=True, size=14)

headers = ['Crew/Promoter', 'Their Gear', 'Deal Type', 'Affiliate %', 'Status', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Empty rows for tracking
for row in range(4, 12):
    for col in range(1, 7):
        ws5.cell(row=row, column=col).border = thin_border

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 12
ws5.column_dimensions['F'].width = 30

# Sheet 6: Vendor Tracker
ws6 = wb.create_sheet("Vendors")

ws6['A1'] = 'VENDOR TRACKER'
ws6['A1'].font = Font(bold=True, size=14)

headers = ['Vendor Name', 'Category', 'Fee Type', 'Fee Amount', 'Contact', 'Status']
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Empty rows for tracking
for row in range(4, 15):
    for col in range(1, 7):
        ws6.cell(row=row, column=col).border = thin_border

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 18
ws6.column_dimensions['C'].width = 12
ws6.column_dimensions['D'].width = 12
ws6.column_dimensions['E'].width = 20
ws6.column_dimensions['F'].width = 12

# Sheet 7: Compensation Model
ws7 = wb.create_sheet("Compensation")

ws7['A1'] = 'COMPENSATION MODEL'
ws7['A1'].font = Font(bold=True, size=14)

# Profit Pool Structure
ws7['A3'] = 'PROFIT POOL - TIER CONTRIBUTIONS'
ws7['A3'].font = Font(bold=True)

headers = ['Phase', 'Per Tier Reached', 'Tiers', 'Phase Total']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

tier_contribs = [
    ('Phase 1 (Tier 1-3)', 300, 3, 900),
    ('Phase 2 (Tier 4-7)', 500, 4, 2000),
    ('Phase 3 (Tier 8-11)', 750, 4, 3000),
    ('Phase 4 (Tier 12-15)', 1000, 4, 4000),
]

for row, (phase, per_tier, tiers, total) in enumerate(tier_contribs, 5):
    ws7.cell(row=row, column=1, value=phase).border = thin_border
    c = ws7.cell(row=row, column=2, value=per_tier)
    c.number_format = currency_format
    c.border = thin_border
    ws7.cell(row=row, column=3, value=tiers).border = thin_border
    c = ws7.cell(row=row, column=4, value=total)
    c.number_format = currency_format
    c.border = thin_border

ws7['A9'] = 'MAX FROM ALL TIERS:'
ws7['A9'].font = Font(bold=True)
ws7['D9'] = 9900
ws7['D9'].number_format = currency_format
ws7['D9'].font = Font(bold=True)

# Affiliate Links
ws7['A12'] = 'INDIVIDUAL AFFILIATE LINKS'
ws7['A12'].font = Font(bold=True)

headers = ['Role', 'Per Ticket Sold', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=13, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

affiliates = [
    ('Stage Partners', '$3-4', 'Highest rate - bringing production'),
    ('Local/Opener DJs', '$2', 'Playing free, earn through promo'),
    ('Vendors', '$2', 'Reach their customer base'),
    ('Contracted Artists', '$1', 'Optional - already paid'),
]

for row, (role, rate, notes) in enumerate(affiliates, 14):
    ws7.cell(row=row, column=1, value=role).border = thin_border
    ws7.cell(row=row, column=2, value=rate).border = thin_border
    ws7.cell(row=row, column=3, value=notes).border = thin_border

# Who Gets What
ws7['A20'] = 'WHO GETS PAID WHAT'
ws7['A20'].font = Font(bold=True)

headers = ['Role', 'Upfront', 'Affiliate', 'Profit Pool']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=21, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

roles = [
    ('Headliners', 'Contract', '$1/tix (opt)', 'No'),
    ('Support Acts', 'Contract', '$1/tix (opt)', 'No'),
    ('Local/Opener DJs', 'Free', '$2/ticket', 'YES'),
    ('Vendors', 'Booth fee', '$2/ticket', 'No'),
    ('Subphonic', '$3-6k', 'No', 'No'),
    ('Keith (Dome)', '$900', 'No', 'No'),
    ('Stage Partners', 'Base (Ph3+)', '$3-4/ticket', 'YES'),
    ('Volunteers', 'Pass+meals', 'No', 'No'),
    ('Organizers', '$0', 'No', 'YES'),
]

for row, (role, upfront, aff, pool) in enumerate(roles, 22):
    ws7.cell(row=row, column=1, value=role).border = thin_border
    ws7.cell(row=row, column=2, value=upfront).border = thin_border
    ws7.cell(row=row, column=3, value=aff).border = thin_border
    c = ws7.cell(row=row, column=4, value=pool)
    c.border = thin_border
    if pool == 'YES':
        c.font = Font(bold=True, color='2B5797')

# Profit Pool Distribution
ws7['A33'] = 'PROFIT POOL DISTRIBUTION'
ws7['A33'].font = Font(bold=True)

ws7['A34'] = 'Phase 1-2 (No Partners)'
ws7['A34'].font = Font(italic=True)

headers = ['Recipient', '% of Pool']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=35, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

dist_p12 = [
    ('Local/Opener Pool', '30%'),
    ('Organizers', '50%'),
    ('Next Year Seed', '20%'),
]

for row, (recip, pct) in enumerate(dist_p12, 36):
    ws7.cell(row=row, column=1, value=recip).border = thin_border
    ws7.cell(row=row, column=2, value=pct).border = thin_border

ws7['A40'] = 'Phase 3-4 (With Partners)'
ws7['A40'].font = Font(italic=True)

for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=41, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

dist_p34 = [
    ('Stage Partners (split)', '20%'),
    ('Local/Opener Pool', '25%'),
    ('Organizers', '40%'),
    ('Next Year Seed', '15%'),
]

for row, (recip, pct) in enumerate(dist_p34, 42):
    ws7.cell(row=row, column=1, value=recip).border = thin_border
    ws7.cell(row=row, column=2, value=pct).border = thin_border

# Profit Pool Calculator
ws7['A48'] = 'PROFIT POOL CALCULATOR'
ws7['A48'].font = Font(bold=True)

ws7['A49'] = 'Highest tier reached:'
ws7['B49'] = 0
ws7['B49'].font = input_font

ws7['A50'] = 'Remaining surplus:'
ws7['B50'] = 0
ws7['B50'].font = input_font
ws7['B50'].number_format = currency_format

ws7['A52'] = 'Tier contributions:'
ws7['B52'] = '=IF(B49>=12,9900,IF(B49>=8,5900,IF(B49>=4,2900,IF(B49>=1,B49*300,0))))'
ws7['B52'].number_format = currency_format

ws7['A53'] = 'Total Profit Pool:'
ws7['B53'] = '=B52+B50'
ws7['B53'].number_format = currency_format
ws7['B53'].font = Font(bold=True)

# Local DJ Pool Calculator
ws7['A56'] = 'LOCAL DJ POOL (30% of profit pool in Ph1-2, 25% in Ph3-4)'
ws7['A56'].font = Font(bold=True)

headers = ['Set Length', 'Points', 'Count', 'Total Pts', 'Per DJ Payout']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=57, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

sets = [
    ('1 hour', 2),
    ('45 min', 1.5),
    ('30 min', 1),
]

for row, (length, pts) in enumerate(sets, 58):
    ws7.cell(row=row, column=1, value=length).border = thin_border
    ws7.cell(row=row, column=2, value=pts).border = thin_border
    count_cell = ws7.cell(row=row, column=3, value=0)
    count_cell.font = input_font
    count_cell.border = thin_border
    ws7.cell(row=row, column=4, value=f'=B{row}*C{row}').border = thin_border
    # Assumes 30% of pool for local artists
    ws7.cell(row=row, column=5, value=f'=IF(SUM($D$58:$D$60)>0,(B{row}/SUM($D$58:$D$60))*($B$53*0.3),0)').border = thin_border
    ws7.cell(row=row, column=5).number_format = currency_format

ws7['A61'] = 'Total Points:'
ws7['D61'] = '=SUM(D58:D60)'
ws7['D61'].font = Font(bold=True)

# Affiliate Calculator
ws7['A64'] = 'AFFILIATE EARNINGS CALCULATOR'
ws7['A64'].font = Font(bold=True)

ws7['A65'] = 'Tickets sold through your link:'
ws7['B65'] = 0
ws7['B65'].font = input_font

ws7['A66'] = 'Your rate ($/ticket):'
ws7['B66'] = 2
ws7['B66'].font = input_font

ws7['A67'] = 'Your affiliate earnings:'
ws7['B67'] = '=B65*B66'
ws7['B67'].number_format = currency_format
ws7['B67'].font = Font(bold=True)

ws7.column_dimensions['A'].width = 30
ws7.column_dimensions['B'].width = 18
ws7.column_dimensions['C'].width = 12
ws7.column_dimensions['D'].width = 12
ws7.column_dimensions['E'].width = 15

wb.save('reeguez_rocks_2026_v3.xlsx')
print("Spreadsheet created successfully")
